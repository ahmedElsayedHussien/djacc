import logging
from django.views.generic import TemplateView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.core.mixins import PermRequiredMixin
from datetime import date
from .models import Supplier, PurchaseInvoice, PurchaseInvoiceLine
from apps.reports.services import ReportService
from django.core.paginator import Paginator
from apps.inventory.models import Item

logger = logging.getLogger(__name__)

class PurchaseReportDashboardView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/dashboard_links.html'
    permission_required = 'purchases.view_purchaseinvoice'

class PurchaseSummaryReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/summary.html'
    permission_required = 'purchases.view_purchaseinvoice'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = self.request.GET.get('from', date.today().replace(day=1).isoformat())
        to_date = self.request.GET.get('to', date.today().isoformat())
        
        from_date_obj = date.fromisoformat(from_date)
        to_date_obj = date.fromisoformat(to_date)
        
        context['report'] = ReportService.purchases_summary_report(from_date_obj, to_date_obj)
        context['from_date'] = from_date_obj
        context['to_date'] = to_date_obj
        return context

class ItemPurchaseCostReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/item_cost.html'
    permission_required = 'purchases.view_purchaseinvoice'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = self.request.GET.get('from', date.today().replace(day=1).isoformat())
        to_date = self.request.GET.get('to', date.today().isoformat())
        
        from_date_obj = date.fromisoformat(from_date)
        to_date_obj = date.fromisoformat(to_date)
        
        report_data = ReportService.item_purchase_cost_report(from_date_obj, to_date_obj)
        paginator = Paginator(report_data, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)
        
        context['from_date'] = from_date_obj
        context['to_date'] = to_date_obj
        return context

class SupplierBalancesReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/supplier_balances.html'
    permission_required = 'purchases.view_supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = self.request.GET.get('from', date.today().replace(day=1).isoformat())
        to_date = self.request.GET.get('to', date.today().isoformat())
        
        from_date_obj = date.fromisoformat(from_date)
        to_date_obj = date.fromisoformat(to_date)
        
        report_data = ReportService.supplier_balances_report(from_date_obj, to_date_obj)
        paginator = Paginator(report_data, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)
        
        context['from_date'] = from_date_obj
        context['to_date'] = to_date_obj
        return context

class SupplierAgingReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/aging.html'
    permission_required = 'purchases.view_supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier_id = self.request.GET.get('supplier')
        
        report_data = ReportService.supplier_aging_report(supplier_id)
        paginator = Paginator(report_data, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)
        
        context['suppliers'] = Supplier.objects.all().select_related('account')
        if supplier_id:
            context['selected_supplier'] = int(supplier_id)
        return context

class OpenPurchaseOrdersReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/open_orders.html'
    permission_required = 'purchases.view_purchaseorder'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['report'] = ReportService.open_purchase_orders_report()
        return context

class PurchaseReturnAnalysisReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/returns_analysis.html'
    permission_required = 'purchases.view_purchasereturn'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = self.request.GET.get('from', date.today().replace(day=1).isoformat())
        to_date = self.request.GET.get('to', date.today().isoformat())
        
        from_date_obj = date.fromisoformat(from_date)
        to_date_obj = date.fromisoformat(to_date)
        
        context['report'] = ReportService.purchase_returns_analysis_report(from_date_obj, to_date_obj)
        context['from_date'] = from_date_obj
        context['to_date'] = to_date_obj
        return context

class ItemPriceFluctuationReportView(LoginRequiredMixin, PermRequiredMixin, TemplateView):
    template_name = 'purchases/reports/item_price_fluctuation.html'
    permission_required = 'purchases.view_purchaseinvoice'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = self.request.GET.get('from', date.today().replace(day=1).isoformat())
        to_date = self.request.GET.get('to', date.today().isoformat())
        item_id = self.request.GET.get('item')
        
        from_date_obj = date.fromisoformat(from_date)
        to_date_obj = date.fromisoformat(to_date)
        
        report_data = []
        if item_id:
            report_data = ReportService.item_price_fluctuation_report(item_id, from_date_obj, to_date_obj)
            context['selected_item'] = int(item_id)
            
        paginator = Paginator(report_data, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)
        
        context['items'] = Item.objects.all().order_by('name')
        context['from_date'] = from_date_obj
        context['to_date'] = to_date_obj
        return context

class DetailedPurchaseReportView(LoginRequiredMixin, PermRequiredMixin, ListView):
    template_name = 'purchases/reports/detailed_purchases.html'
    context_object_name = 'report_data'
    permission_required = 'purchases.view_purchaseinvoice'
    paginate_by = 100

    def get_queryset(self):
        date_from_str = self.request.GET.get('date_from')
        date_to_str = self.request.GET.get('date_to')
        supplier_id = self.request.GET.get('supplier')
        
        qs = PurchaseInvoiceLine.objects.filter(invoice__status=PurchaseInvoice.Status.POSTED).select_related(
            'invoice', 'invoice__supplier', 'item'
        ).order_by('-invoice__date', '-invoice__number')
        
        if date_from_str:
            qs = qs.filter(invoice__date__gte=date.fromisoformat(date_from_str))
        if date_to_str:
            qs = qs.filter(invoice__date__lte=date.fromisoformat(date_to_str))
        if supplier_id:
            qs = qs.filter(invoice__supplier_id=supplier_id)
            
        return qs

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font
            from django.http import HttpResponse
            
            queryset = self.get_queryset()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detailed Purchases"
            
            # Set sheet direction to RTL for Arabic
            ws.views.sheetView[0].showGridLines = True
            
            headers = ["التاريخ", "رقم الفاتورة", "المورد", "الصنف", "الكمية", "سعر الشراء", "الخصم", "الضريبة", "الصافي"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True)
                
            for row_idx, line in enumerate(queryset, 2):
                ws.cell(row=row_idx, column=1, value=line.invoice.date.strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=2, value=line.invoice.number)
                ws.cell(row=row_idx, column=3, value=line.invoice.supplier.name)
                ws.cell(row=row_idx, column=4, value=line.item.name)
                ws.cell(row=row_idx, column=5, value=float(line.quantity))
                ws.cell(row=row_idx, column=6, value=float(line.unit_cost))
                ws.cell(row=row_idx, column=7, value=float(line.discount_amount))
                ws.cell(row=row_idx, column=8, value=float(line.tax_amount))
                ws.cell(row=row_idx, column=9, value=float(line.total))
                
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="detailed_purchases.xlsx"'
            wb.save(response)
            return response
            
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['suppliers'] = Supplier.objects.all()
        return context
