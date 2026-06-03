import logging
import openpyxl
from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal
from openpyxl.styles import Font

from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.core.exceptions import PermissionDenied

from .services import ReportService
from apps.core.models import FiscalYear, CostCenter, Account
from apps.sales.models import Customer, SalesRepresentative, SalesInvoice, SalesTarget, CustomerReceipt, RepDailySettlement, SalesReturn
from apps.purchases.models import Supplier
from apps.inventory.models import Item, Warehouse, ItemLedger, StockVoucher

logger = logging.getLogger(__name__)

def safe_int(val, default=None):
    if not val:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

def parse_date(date_val, default=None):
    if not date_val:
        return default or date.today()
    if isinstance(date_val, str):
        try:
            return datetime.strptime(date_val, '%Y-%m-%d').date()
        except ValueError:
            return default or date.today()
    return date_val

class ExcelExportMixin:
    excel_filename = 'report.xlsx'

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get('export') == 'xlsx':
            return self._generate_excel(context)
        return super().render_to_response(context, **response_kwargs)

    def _generate_excel(self, context):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        columns = self.get_excel_columns()
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col[0])
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(self.get_excel_rows(context), 2):
            for col_idx, col in enumerate(columns, 1):
                val = col[1](row)
                if hasattr(val, 'as_tuple'):
                    val = float(val)
                elif val is None:
                    val = ''
                ws.cell(row=row_idx, column=col_idx, value=val)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.excel_filename}"'
        wb.save(response)
        return response

    def get_excel_columns(self):
        return []

    def get_excel_rows(self, context):
        return []


class FinancialReportDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/financial_dashboard.html'


class TrialBalanceView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/trial_balance.html'
    excel_filename = 'trial_balance.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date(date.today().year, 1, 1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['cost_centers'] = CostCenter.objects.filter(is_active=True)
        cc_id = safe_int(self.request.GET.get('cost_center'))
        if cc_id:
            context['selected_cc'] = cc_id

        rows = ReportService.trial_balance(from_date, to_date)
        context['rows'] = rows

        totals = {
            'op_debit': sum(row['op_debit'] for row in rows),
            'op_credit': sum(row['op_credit'] for row in rows),
            'mov_debit': sum(row['mov_debit'] for row in rows),
            'mov_credit': sum(row['mov_credit'] for row in rows),
            'cl_debit': sum(row['cl_debit'] for row in rows),
            'cl_credit': sum(row['cl_credit'] for row in rows),
        }
        context['totals'] = totals

        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('كود الحساب', lambda r: r['account'].code),
            ('اسم الحساب', lambda r: r['account'].name),
            ('رصيد افتتاحي مدين', lambda r: r['op_debit']),
            ('رصيد افتتاحي دائن', lambda r: r['op_credit']),
            ('حركة الفترة مدين', lambda r: r['mov_debit']),
            ('حركة الفترة دائن', lambda r: r['mov_credit']),
            ('رصيد ختامي مدين', lambda r: r['cl_debit']),
            ('رصيد ختامي دائن', lambda r: r['cl_credit']),
        ]

    def get_excel_rows(self, context):
        return context['rows']


class IncomeStatementView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/income_statement.html'
    excel_filename = 'income_statement.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['cost_centers'] = CostCenter.objects.filter(is_active=True)
        cc_id_int = safe_int(self.request.GET.get('cost_center'))
        if cc_id_int:
            context['selected_cc'] = cc_id_int

        context['report'] = ReportService.income_statement(from_date, to_date, cost_center_id=cc_id_int)
        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('البيان', lambda r: r[0]),
            ('القيمة', lambda r: r[1]),
        ]

    def get_excel_rows(self, context):
        r = context['report']
        rows = []
        rows.append(('إيرادات المبيعات', float(r['sales'])))
        rows.append(('مردودات المبيعات', float(r['sales_returns'])))
        rows.append(('خصم المبيعات', float(r['sales_discount'])))
        rows.append(('صافي المبيعات', float(r['net_sales'])))
        rows.append(('', ''))
        rows.append(('تكلفة المبيعات', ''))
        for item in r['cogs_items']:
            rows.append((f"  {item['name']}", float(item['balance'])))
        rows.append(('إجمالي تكلفة المبيعات', float(r['cogs_total'])))
        rows.append(('', ''))
        rows.append(('إجمالي الربح', float(r['gross_profit'])))
        rows.append(('', ''))
        rows.append(('مصروفات التشغيل', ''))
        for item in r['op_expenses_items']:
            rows.append((f"  {item['name']}", float(item['balance'])))
        rows.append(('إجمالي مصروفات التشغيل', float(r['total_op_expenses'])))
        rows.append(('', ''))
        rows.append(('ربح التشغيل', float(r['operating_profit'])))
        rows.append(('', ''))
        rows.append(('إيرادات أخرى', ''))
        for item in r['other_rev_items']:
            rows.append((f"  {item['name']}", float(item['balance'])))
        rows.append(('إجمالي الإيرادات الأخرى', float(r['total_other_rev'])))
        rows.append(('مصروفات تمويلية', ''))
        for item in r['finance_exp_items']:
            rows.append((f"  {item['name']}", float(item['balance'])))
        rows.append(('إجمالي المصروفات التمويلية', float(r['total_finance_exp'])))
        rows.append(('مصروفات أخرى', ''))
        for item in r['other_exp_items']:
            rows.append((f"  {item['name']}", float(item['balance'])))
        rows.append(('إجمالي المصروفات الأخرى', float(r['total_other_exp'])))
        rows.append(('', ''))
        rows.append(('صافي الربح قبل الضريبة', float(r['net_profit_before_tax'])))
        rows.append(('مصروف الضريبة', float(r['tax_exp'])))
        rows.append(('', ''))
        rows.append(('صافي الدخل', float(r['net_income'])))
        return rows


class BalanceSheetView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/balance_sheet.html'
    excel_filename = 'balance_sheet.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        as_of_date = parse_date(self.request.GET.get('date'), date.today())

        context['cost_centers'] = CostCenter.objects.filter(is_active=True)
        cc_id_int = safe_int(self.request.GET.get('cost_center'))
        if cc_id_int:
            context['selected_cc'] = cc_id_int

        context['report'] = ReportService.balance_sheet(as_of_date, cost_center_id=cc_id_int)
        context['as_of_date'] = as_of_date
        return context

    def get_excel_columns(self):
        return [
            ('البيان', lambda r: r[0]),
            ('القيمة', lambda r: r[1]),
        ]

    def get_excel_rows(self, context):
        r = context['report']
        rows = []
        rows.append(('الأصول', ''))
        for a in r['assets']:
            rows.append((f"  {a['name']}", float(a['balance'])))
        rows.append(('إجمالي الأصول', float(r['total_assets'])))
        rows.append(('', ''))
        rows.append(('الخصوم', ''))
        for l in r['liabilities']:
            rows.append((f"  {l['name']}", float(l['balance'])))
        rows.append(('إجمالي الخصوم', float(r['total_liabilities'])))
        rows.append(('', ''))
        rows.append(('حقوق الملكية', ''))
        for e in r['equity']:
            rows.append((f"  {e['name']}", float(e['balance'])))
        rows.append(('إجمالي حقوق الملكية', float(r['total_equity'])))
        return rows


class CustomerStatementView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'sales.view_customer'
    template_name = 'reports/customer_statement.html'
    excel_filename = 'customer_statement.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer_id = safe_int(self.request.GET.get('customer'))
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['customers'] = Customer.objects.all()
        context['from_date'] = from_date
        context['to_date'] = to_date

        if customer_id:
            context['report'] = ReportService.customer_statement(customer_id, from_date, to_date)
            context['selected_customer'] = customer_id

        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: str(r['date'])),
            ('رقم المستند', lambda r: r['number']),
            ('البيان', lambda r: r['description']),
            ('مدين', lambda r: r['debit']),
            ('دائن', lambda r: r['credit']),
            ('الرصيد', lambda r: r['balance']),
        ]

    def get_excel_rows(self, context):
        r = context.get('report')
        if not r:
            return []
        opening = r['opening_balance']
        lines = [{'date': '', 'number': '', 'description': 'رصيد افتتاحي', 'debit': 0, 'credit': 0, 'balance': float(opening)}]
        for line in r['lines']:
            lines.append(line)
        return lines


class SupplierStatementView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'purchases.view_supplier'
    template_name = 'reports/supplier_statement.html'
    excel_filename = 'supplier_statement.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier_id = safe_int(self.request.GET.get('supplier'))
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['suppliers'] = Supplier.objects.all()
        context['from_date'] = from_date
        context['to_date'] = to_date

        if supplier_id:
            context['report'] = ReportService.supplier_statement(supplier_id, from_date, to_date)
            context['selected_supplier'] = supplier_id

        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: str(r['date'])),
            ('رقم المستند', lambda r: r.get('number', '')),
            ('البيان', lambda r: r['description']),
            ('مدين', lambda r: r['debit']),
            ('دائن', lambda r: r['credit']),
            ('الرصيد', lambda r: r['balance']),
        ]

    def get_excel_rows(self, context):
        r = context.get('report')
        if not r:
            return []
        op = r['op_balance']
        lines = [{'date': '', 'number': '', 'description': 'رصيد افتتاحي', 'debit': 0, 'credit': 0, 'balance': float(op)}]
        for m in r['movements']:
            lines.append(m)
        return lines


class RepStatementView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'sales.view_salesrepresentative'
    template_name = 'reports/rep_statement.html'
    excel_filename = 'rep_statement.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rep_id = safe_int(self.request.GET.get('rep'))
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['reps'] = SalesRepresentative.objects.all()
        context['from_date'] = from_date
        context['to_date'] = to_date

        if rep_id:
            context['report'] = ReportService.rep_statement(rep_id, from_date, to_date)
            context['selected_rep'] = rep_id

        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: str(r['date'])),
            ('البيان', lambda r: r['description']),
            ('مدين', lambda r: r['debit']),
            ('دائن', lambda r: r['credit']),
            ('الرصيد', lambda r: r['balance']),
        ]

    def get_excel_rows(self, context):
        r = context.get('report')
        if not r:
            return []
        op = r['opening_balance']
        lines = [{'date': '', 'description': 'رصيد افتتاحي', 'debit': 0, 'credit': 0, 'balance': float(op)}]
        for line in r['lines']:
            lines.append(line)
        return lines


class StockStatusView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_item'
    template_name = 'reports/stock_status.html'
    excel_filename = 'stock_status.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        warehouse_id = safe_int(self.request.GET.get('warehouse'))

        context['warehouses'] = Warehouse.objects.all()
        context['_full_report'] = ReportService.stock_status(warehouse_id)
        report_data = context['_full_report']

        paginator = Paginator(report_data['items'], 50)
        page_number = self.request.GET.get('page')
        report_data['items'] = paginator.get_page(page_number)

        context['report'] = report_data
        if warehouse_id:
            context['selected_warehouse'] = warehouse_id

        return context

    def get_excel_columns(self):
        return [
            ('المخزن', lambda r: r.warehouse.name),
            ('الصنف', lambda r: r.item.name),
            ('الوحدة', lambda r: r.item.base_unit.symbol if r.item.base_unit else ''),
            ('الكمية', lambda r: r.quantity_on_hand),
            ('سعر التكلفة', lambda r: r.cost_price),
            ('إجمالي القيمة', lambda r: r.total_value),
        ]

    def get_excel_rows(self, context):
        full = context.get('_full_report')
        if not full:
            return []
        return list(full['items'])


class RepCommissionView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'sales.view_salesrepresentative'
    template_name = 'reports/rep_commission.html'
    excel_filename = 'rep_commission.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())
        context['report'] = ReportService.rep_commission_report(from_date, to_date)
        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('المندوب', lambda r: r['rep'].name),
            ('إجمالي المبيعات', lambda r: r['total_sales']),
            ('نسبة العمولة', lambda r: float(r['commission_rate'])),
            ('قيمة العمولة', lambda r: r['commission_amount']),
        ]

    def get_excel_rows(self, context):
        return context['report']


class CostCenterStatementView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_costcenter'
    template_name = 'reports/cost_center_statement.html'
    excel_filename = 'cost_center_statement.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cost_center_id = safe_int(self.request.GET.get('cost_center'))
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['cost_centers'] = CostCenter.objects.all()
        context['from_date'] = from_date
        context['to_date'] = to_date

        if cost_center_id:
            context['report'] = ReportService.cost_center_statement(cost_center_id, from_date, to_date)
            context['selected_cost_center'] = cost_center_id

        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: str(r['date'])),
            ('الحساب', lambda r: r['account']),
            ('البيان', lambda r: r['description']),
            ('مدين', lambda r: r['debit']),
            ('دائن', lambda r: r['credit']),
        ]

    def get_excel_rows(self, context):
        r = context.get('report')
        if not r:
            return []
        return r['lines']


class AccountStatementView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/account_statement.html'
    excel_filename = 'account_statement.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account_id = safe_int(self.request.GET.get('account'))
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['accounts'] = Account.objects.filter(is_leaf=True).order_by('code')
        context['from_date'] = from_date
        context['to_date'] = to_date

        if account_id:
            context['report'] = ReportService.account_statement(account_id, from_date, to_date)
            context['selected_account'] = account_id

        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: str(r['date'])),
            ('رقم القيد', lambda r: r['entry_number']),
            ('البيان', lambda r: r['description']),
            ('مدين', lambda r: r['debit']),
            ('دائن', lambda r: r['credit']),
            ('الرصيد', lambda r: r['balance']),
        ]

    def get_excel_rows(self, context):
        r = context.get('report')
        if not r:
            return []
        op = r['opening_balance']
        lines = [{'date': '', 'entry_number': '', 'description': 'رصيد افتتاحي', 'debit': 0, 'credit': 0, 'balance': float(op)}]
        for line in r['lines']:
            lines.append(line)
        return lines


class VATReportView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/vat_report.html'
    excel_filename = 'vat_report.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['cost_centers'] = CostCenter.objects.filter(is_active=True)
        cc_id_int = safe_int(self.request.GET.get('cost_center'))
        if cc_id_int:
            context['selected_cc'] = cc_id_int

        context['report'] = ReportService.vat_report(from_date, to_date, cost_center_id=cc_id_int)
        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('البيان', lambda r: r[0]),
            ('القيمة', lambda r: r[1]),
        ]

    def get_excel_rows(self, context):
        r = context['report']
        return [
            ('ضريبة المبيعات (Output VAT)', float(r['output_vat'])),
            ('عدد فواتير المبيعات', r['output_vat_count']),
            ('ضريبة المشتريات (Input VAT)', float(r['input_vat'])),
            ('عدد فواتير المشتريات', r['input_vat_count']),
            ('صافي ضريبة القيمة المضافة', float(r['net_vat'])),
            ('الحالة', 'مستحق للضرائب' if r['is_payable'] else 'مدين للشركة'),
        ]


class WHTReportView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'core.view_account'
    template_name = 'reports/wht_report.html'
    excel_filename = 'wht_report.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['cost_centers'] = CostCenter.objects.filter(is_active=True)
        cc_id_int = safe_int(self.request.GET.get('cost_center'))
        if cc_id_int:
            context['selected_cc'] = cc_id_int

        context['report'] = ReportService.wht_report(from_date, to_date, cost_center_id=cc_id_int)
        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('البيان', lambda r: r[0]),
            ('القيمة', lambda r: r[1]),
        ]

    def get_excel_rows(self, context):
        r = context['report']
        return [
            ('WHT على المبيعات', float(r['wht_on_sales'])),
            ('حساب WHT المبيعات', r['wht_on_sales_account']),
            ('WHT على المشتريات', float(r['wht_on_purchases'])),
            ('حساب WHT المشتريات', r['wht_on_purchases_account']),
            ('صافي WHT', float(r['net_wht'])),
            ('الحالة', 'مستحق من الضرائب' if r['is_receivable'] else 'مستحق للضرائب'),
        ]



class InventoryValuationView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_item'
    template_name = 'reports/inventory_valuation.html'
    excel_filename = 'inventory_valuation.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        warehouse_id = safe_int(self.request.GET.get('warehouse'))
        context['warehouses'] = Warehouse.objects.all()
        context['_full_report'] = ReportService.inventory_valuation(warehouse_id)
        report_data = context['_full_report']

        paginator = Paginator(report_data['items'], 50)
        page_number = self.request.GET.get('page')
        report_data['items'] = paginator.get_page(page_number)

        context['report'] = report_data
        if warehouse_id:
            context['selected_warehouse'] = warehouse_id

        return context

    def get_excel_columns(self):
        return [
            ('المخزن', lambda r: r.warehouse.name),
            ('التصنيف', lambda r: r.item.category.name if r.item.category else ''),
            ('الصنف', lambda r: r.item.name),
            ('الكمية', lambda r: r.quantity_on_hand),
            ('سعر التكلفة', lambda r: r.cost_price),
            ('إجمالي القيمة', lambda r: r.total_value),
        ]

    def get_excel_rows(self, context):
        full = context.get('_full_report')
        if not full:
            return []
        return list(full['items'])


class ReorderAlertView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_item'
    template_name = 'reports/reorder_alert.html'
    excel_filename = 'reorder_alert.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['_full_qs'] = ReportService.reorder_alert_report()
        report_qs = context['_full_qs']

        paginator = Paginator(report_qs, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)

        return context

    def get_excel_columns(self):
        return [
            ('كود الصنف', lambda r: r.code),
            ('اسم الصنف', lambda r: r.name),
            ('المخزون الحالي', lambda r: float(r.current_stock)),
            ('الحد الأدنى', lambda r: float(r.minimum_stock)),
        ]

    def get_excel_rows(self, context):
        full = context.get('_full_qs')
        if full is None:
            return []
        return list(full)


class ItemLedgerReportView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_item'
    template_name = 'reports/item_ledger.html'
    excel_filename = 'item_ledger.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item_id = safe_int(self.request.GET.get('item'))
        warehouse_id = safe_int(self.request.GET.get('warehouse'))
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())

        context['items'] = Item.objects.filter(is_active=True).order_by('name')
        context['warehouses'] = Warehouse.objects.all()
        context['from_date'] = from_date
        context['to_date'] = to_date

        if item_id:
            context['_full_report'] = ReportService.item_ledger_report(
                item_id,
                warehouse_id,
                from_date,
                to_date
            )
            report_data = context['_full_report']

            paginator = Paginator(report_data['movements'], 50)
            page_number = self.request.GET.get('page')
            report_data['movements'] = paginator.get_page(page_number)

            context['report'] = report_data
            context['selected_item'] = item_id
            if warehouse_id:
                context['selected_warehouse'] = warehouse_id

        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: r.get('date', '')),
            ('المخزن', lambda r: r.get('warehouse', '')),
            ('البيان', lambda r: r.get('description', '')),
            ('الكمية', lambda r: float(r.get('quantity', 0))),
            ('الرصيد', lambda r: float(r.get('running_balance', 0))),
        ]

    def get_excel_rows(self, context):
        r = context.get('_full_report')
        if not r:
            return []
        op = r['opening_qty']
        rows = [dict(date='', warehouse='', description='رصيد افتتاحي', quantity=0, running_balance=float(op))]
        for m in r['movements']:
            rows.append(dict(
                date=str(getattr(m, 'date', '')),
                warehouse=m.warehouse.name if hasattr(m, 'warehouse') and m.warehouse else '',
                description=m.description if hasattr(m, 'description') else getattr(m, 'reference', ''),
                quantity=float(getattr(m, 'quantity', 0)),
                running_balance=0,
            ))
        bal = float(op)
        for row in rows[1:]:
            bal += row['quantity']
            row['running_balance'] = bal
        return rows


class WastageAdjustmentsView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_stockvoucher'
    template_name = 'reports/wastage_adjustments.html'
    excel_filename = 'wastage_adjustments.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date.today().replace(day=1))
        to_date = parse_date(self.request.GET.get('to'), date.today())
        context['_full_qs'] = ReportService.wastage_adjustments_report(from_date, to_date)
        report_qs = context['_full_qs']

        paginator = Paginator(report_qs, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)

        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('التاريخ', lambda r: str(r.date)),
            ('رقم الإذن', lambda r: r.number),
            ('المستودع', lambda r: r.warehouse.name),
            ('الحالة', lambda r: r.get_status_display()),
            ('نوع التسوية', lambda r: r.get_voucher_type_display()),
            ('الإجمالي', lambda r: float(r.total)),
        ]

    def get_excel_rows(self, context):
        full = context.get('_full_qs')
        if full is None:
            return []
        return list(full)


class VanInventoryView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'sales.view_salesrepresentative'
    template_name = 'reports/van_inventory.html'
    excel_filename = 'van_inventory.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rep_id = safe_int(self.request.GET.get('rep'))
        context['reps'] = SalesRepresentative.objects.filter(is_active=True)

        if rep_id:
            try:
                context['_full_report'] = ReportService.van_inventory_report(rep_id)
                report_data = context['_full_report']

                paginator = Paginator(report_data['items'], 50)
                page_number = self.request.GET.get('page')
                report_data['items'] = paginator.get_page(page_number)

                context['report'] = report_data
                context['selected_rep'] = rep_id
            except Exception as e:
                context['error'] = str(e)

        return context

    def get_excel_columns(self):
        return [
            ('الصنف', lambda r: r.item.name),
            ('التصنيف', lambda r: r.item.category.name if r.item.category else ''),
            ('الكمية', lambda r: r.quantity_on_hand),
            ('سعر التكلفة', lambda r: r.cost_price),
            ('إجمالي القيمة', lambda r: r.total_value),
        ]

    def get_excel_rows(self, context):
        full = context.get('_full_report')
        if not full:
            return []
        return list(full['items'])


class InventoryTurnoverView(ExcelExportMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'inventory.view_item'
    template_name = 'reports/inventory_turnover.html'
    excel_filename = 'inventory_turnover.xlsx'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from_date = parse_date(self.request.GET.get('from'), date(date.today().year, 1, 1))
        to_date = parse_date(self.request.GET.get('to'), date.today())
        context['_full_list'] = ReportService.inventory_turnover_report(from_date, to_date)
        report_list = context['_full_list']

        paginator = Paginator(report_list, 50)
        page_number = self.request.GET.get('page')
        context['report'] = paginator.get_page(page_number)

        context['from_date'] = from_date
        context['to_date'] = to_date
        return context

    def get_excel_columns(self):
        return [
            ('كود الصنف', lambda r: r['item'].code),
            ('اسم الصنف', lambda r: r['item'].name),
            ('تكلفة المبيعات', lambda r: float(r['cogs'])),
            ('متوسط المخزون', lambda r: float(r['avg_inventory'])),
            ('نسبة الدوران', lambda r: float(r['turnover_ratio'])),
        ]

    def get_excel_rows(self, context):
        full = context.get('_full_list')
        if full is None:
            return []
        return full


class SalesRepDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    permission_required = 'sales.view_salesrepresentative'
    template_name = 'reports/rep_dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        is_rep = hasattr(request.user, 'salesrepresentative')
        is_admin = request.user.is_superuser or request.user.is_staff
        if not (is_rep or is_admin):
            raise PermissionDenied("ليس لديك صلاحية للوصول إلى هذه لوحة التحكم.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user
        rep = None

        if hasattr(user, 'salesrepresentative'):
            rep = user.salesrepresentative
            context['is_representative'] = True
        else:
            rep_id = self.request.GET.get('rep')
            if rep_id:
                try:
                    rep = SalesRepresentative.objects.get(id=int(rep_id))
                except (ValueError, SalesRepresentative.DoesNotExist):
                    pass
            context['is_representative'] = False
            context['reps'] = SalesRepresentative.objects.filter(is_active=True)

        context['selected_rep'] = rep

        if rep:
            today = date.today()
            start_of_month = date(today.year, today.month, 1)

            selected_date_str = self.request.GET.get('date')
            selected_date = today
            if selected_date_str:
                try:
                    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
                except ValueError:
                    logger.warning("تاريخ غير صالح في تقرير مندوب: %s", selected_date_str)
            context['selected_date'] = selected_date
            context['selected_date_str'] = selected_date_str

            safe_balance = rep.cash_box.current_balance if rep.cash_box else Decimal('0')
            context['safe_balance'] = safe_balance

            van_items = ItemLedger.objects.filter(warehouse=rep.warehouse).select_related('item', 'item__category').order_by('item__name')
            van_inventory_value = van_items.aggregate(total=Sum('total_value'))['total'] or Decimal('0')
            context['van_items'] = van_items
            context['van_inventory_value'] = van_inventory_value

            sales_today = SalesInvoice.objects.filter(
                sales_rep=rep,
                status=SalesInvoice.Status.POSTED,
                date=today
            ).aggregate(total=Sum('total'))['total'] or Decimal('0')
            context['sales_today'] = sales_today

            sales_this_month = SalesInvoice.objects.filter(
                sales_rep=rep,
                status=SalesInvoice.Status.POSTED,
                date__range=[start_of_month, today]
            ).aggregate(total=Sum('total'))['total'] or Decimal('0')
            context['sales_this_month'] = sales_this_month

            returns_today = SalesReturn.objects.filter(
                sales_rep=rep,
                status=SalesReturn.Status.POSTED,
                date=today
            ).aggregate(total=Sum('total'))['total'] or Decimal('0')
            context['returns_today'] = returns_today

            returns_this_month = SalesReturn.objects.filter(
                sales_rep=rep,
                status=SalesReturn.Status.POSTED,
                date__range=[start_of_month, today]
            ).aggregate(total=Sum('total'))['total'] or Decimal('0')
            context['returns_this_month'] = returns_this_month

            target = SalesTarget.objects.filter(
                sales_rep=rep,
                start_date__lte=today,
                end_date__gte=today
            ).first()

            context['target'] = target
            if target and target.target_amount > 0:
                context['target_amount'] = target.target_amount
                context['target_achievement'] = min(int(float(sales_this_month / target.target_amount) * 100), 100)
            else:
                context['target_amount'] = Decimal('0')
                context['target_achievement'] = None

            commission_rate = rep.commission_rate or Decimal('0')
            context['commission_rate'] = commission_rate
            context['commission_earned'] = (sales_this_month * commission_rate) / Decimal('100')

            if rep.cash_box and rep.cash_box.account:
                cash_box_statement = ReportService.account_statement(rep.cash_box.account.id, selected_date, selected_date)
                context['cash_box_statement'] = cash_box_statement

            recent_invoices = SalesInvoice.objects.filter(sales_rep=rep).order_by('-date', '-id')[:10]
            context['recent_invoices'] = recent_invoices

            recent_returns = SalesReturn.objects.filter(sales_rep=rep).order_by('-date', '-id')[:10]
            context['recent_returns'] = recent_returns

            if rep.cash_box:
                recent_receipts = CustomerReceipt.objects.filter(cash_box=rep.cash_box).order_by('-date', '-id')[:10]
                context['recent_receipts'] = recent_receipts

            recent_settlements = RepDailySettlement.objects.filter(sales_rep=rep).order_by('-date', '-id')[:10]
            context['recent_settlements'] = recent_settlements

            # --- Tab: Invoices & Returns Filter ---
            inv_start_date_str = self.request.GET.get('inv_start_date')
            inv_end_date_str = self.request.GET.get('inv_end_date')
            inv_search = self.request.GET.get('inv_search', '')

            inv_start_date = start_of_month
            inv_end_date = today

            if inv_start_date_str:
                try:
                    inv_start_date = datetime.strptime(inv_start_date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            if inv_end_date_str:
                try:
                    inv_end_date = datetime.strptime(inv_end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass

            context['inv_start_date'] = inv_start_date.strftime('%Y-%m-%d')
            context['inv_end_date'] = inv_end_date.strftime('%Y-%m-%d')
            context['inv_search'] = inv_search

            filtered_invoices = SalesInvoice.objects.filter(
                sales_rep=rep,
                date__range=[inv_start_date, inv_end_date]
            )
            filtered_returns = SalesReturn.objects.filter(
                sales_rep=rep,
                date__range=[inv_start_date, inv_end_date]
            )

            if inv_search:
                from django.db.models import Q
                filtered_invoices = filtered_invoices.filter(
                    Q(number__icontains=inv_search) | 
                    Q(customer__name__icontains=inv_search)
                )
                filtered_returns = filtered_returns.filter(
                    Q(number__icontains=inv_search) | 
                    Q(customer__name__icontains=inv_search)
                )

            context['filtered_invoices'] = filtered_invoices.order_by('-date', '-id')
            context['filtered_returns'] = filtered_returns.order_by('-date', '-id')
            
            context['active_tab'] = self.request.GET.get('tab', 'stock')

        return context
